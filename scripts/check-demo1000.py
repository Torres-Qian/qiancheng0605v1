import openpyxl

wb = openpyxl.load_workbook(r'd:\kaoshi\new\qiancheng0605v1\test-data\demo_1000.xlsx', read_only=True)
ws = wb.active
print(f'Total rows: {ws.max_row} (incl header)')
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print(f'Headers: {headers}')

skus = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[5]:
        skus[row[5]] = skus.get(row[5], 0) + 1
print(f'Unique SKUs: {len(skus)}')
bad = [s for s in skus if 'BAD' in s]
print(f'BAD SKUs count: {len(bad)}')
print(f'BAD examples: {bad[:5]}')
valid = [s for s in skus if 'BAD' not in s]
print(f'Valid SKUs count: {len(valid)}')
print(f'Valid examples: {valid[:5]}')