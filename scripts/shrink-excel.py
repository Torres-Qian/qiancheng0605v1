import openpyxl

wb = openpyxl.load_workbook(r'd:\kaoshi\new\qiancheng0605v1\test-data\10000-orders.xlsx', read_only=True)
ws = wb.active

# 读取表头
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print(f'Headers: {headers}')
print(f'Total rows: {ws.max_row} (incl header)')

# 统计行数和文件大小
rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
total = len(rows)
print(f'Data rows: {total}')

# 砍掉 1000 行，变成 9000 行（文件约 4.0MB，在 4.5MB 以内）
target = 9000
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.append(headers)
for row in rows[:target]:
    out_ws.append(list(row))

output_path = r'd:\kaoshi\new\qiancheng0605v1\test-data\9000-orders.xlsx'
out_wb.save(output_path)

import os
size = os.path.getsize(output_path)
print(f'\nSaved: {output_path}')
print(f'Rows: {target + 1} (incl header)')
print(f'Size: {size:,} bytes ({size/1024/1024:.2f} MB)')
print(f'Within 4.5MB: {"YES" if size < 4.5*1024*1024 else "NO"}')