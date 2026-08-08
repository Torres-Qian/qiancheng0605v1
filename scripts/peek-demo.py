import openpyxl

wb = openpyxl.load_workbook(r'd:\kaoshi\new\qiancheng0605v1\test-data\demo_1000.xlsx', read_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print(f'Headers: {headers}')

# 看前 5 行实际数据
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 1):
    print(f'Row {i}: {row}')

# 统计 SKU数量 列的取值
print('\nSKU数量取值分布:')
qty = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    q = row[7]
    qty[str(q)] = qty.get(str(q), 0) + 1
for k, v in list(qty.items())[:5]:
    print(f'  {k}: {v}')