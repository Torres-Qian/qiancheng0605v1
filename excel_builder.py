"""
Excel Builder — openpyxl 无中间文件方案
从 stdin 读取 JSON 数据，通过 stdout 输出 base64 编码的 .xlsx
供 Node.js child_process 调用，不产生任何持久化文件
"""

import sys
import json
import base64
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# XML 1.0 合法字符：
# #x9 | #xA | #xD | [#x20-#xD7FF] | [#xE000-#xFFFD] | [#x10000-#x10FFFF]
# 即：移除 0x00-0x08, 0x0B, 0x0C, 0x0E-0x1F, 0xD800-0xDFFF, 0xFFFE-0xFFFF
def sanitize_xml(text: str) -> str:
    """移除 XML/Excel 非法字符"""
    result = []
    for ch in text:
        cp = ord(ch)
        if (
            cp == 0x9 or cp == 0xA or cp == 0xD
            or (0x20 <= cp <= 0xD7FF)
            or (0xE000 <= cp <= 0xFFFD)
            or (0x10000 <= cp <= 0x10FFFF)
        ):
            result.append(ch)
    return ''.join(result)


TABLE_HEADER_KEYWORDS = [
    "\u7269\u54c1\u7f16\u7801",    # 物品编码
    "\u7269\u54c1\u540d\u79f0",    # 物品名称
    "\u89c4\u683c",               # 规格
    "\u6570\u91cf",               # 数量
    "\u5355\u4f4d",               # 单位
    "\u5907\u6ce8",               # 备注
]


def is_header_row(row: list) -> bool:
    text = "".join(str(c) for c in row if c)
    return sum(1 for k in TABLE_HEADER_KEYWORDS if k in text) >= 2


def build_excel(rows: list) -> bytes:
    """用 openpyxl 构建格式化 Excel，返回 bytes"""
    if not rows:
        return b""

    n_rows = len(rows)
    n_cols = max(len(r) for r in rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF\u5bfc\u51fa"

    # 写入数据（清洗 XML 非法字符）
    for r_idx, row in enumerate(rows, 1):
        for c_idx in range(1, n_cols + 1):
            val = row[c_idx - 1] if c_idx - 1 < len(row) else ""
            ws.cell(row=r_idx, column=c_idx, value=sanitize_xml(str(val)))

    # 找到表头行
    header_idx = 1
    for i, row in enumerate(rows):
        if is_header_row(row):
            header_idx = i + 1
            break

    # 样式定义
    thin_side = Side(style="thin", color="808080")
    thin_border = Border(
        left=thin_side, right=thin_side,
        top=thin_side, bottom=thin_side,
    )
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # 计算列宽
    for c_idx in range(1, n_cols + 1):
        max_width = 0
        for row in rows:
            val = str(row[c_idx - 1]) if c_idx - 1 < len(row) else ""
            width = sum(
                2 if "\u4e00" <= ch <= "\u9fff" or "\u3000" <= ch <= "\u303f" or "\uff00" <= ch <= "\uffef"
                else 1
                for ch in val
            )
            max_width = max(max_width, width)
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_width + 3, 6), 45)

    # 应用样式
    for r_idx in range(1, n_rows + 1):
        row = rows[r_idx - 1]
        is_header = (r_idx == header_idx)
        is_total = any(
            "\u5408\u8ba1" in str(row[c_idx - 1] if c_idx - 1 < len(row) else "")
            for c_idx in range(1, n_cols + 1)
        )

        for c_idx in range(1, n_cols + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
            cell.border = thin_border
            cell.font = header_font if is_header else normal_font
            if is_header:
                cell.fill = header_fill
            elif is_total:
                cell.fill = total_fill

    # 冻结窗格
    if header_idx <= n_rows:
        ws.freeze_panes = f"A{header_idx + 1}"

    # 自动筛选
    last_col = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{header_idx}:{last_col}{n_rows}"

    # 行高
    ws.row_dimensions[header_idx].height = 22
    for r_idx in range(1, n_rows + 1):
        if r_idx != header_idx:
            ws.row_dimensions[r_idx].height = 18

    # ═══════════════════════════════════════════════════
    # 强制确保最后一列在 XLSX 中可见
    # 某些 XLSX 读取器会忽略空列，导致最后一列（如"备注"）丢失
    # ═══════════════════════════════════════════════════
    last_col_letter = get_column_letter(n_cols)
    last_row = n_rows

    # 1) 在最后一列的最后一行写入占位符，确保维度包含全部列
    last_cell = ws.cell(row=last_row, column=n_cols)
    if last_cell.value is None or str(last_cell.value).strip() == "":
        last_cell.value = " "  # 空格占位，Excel 中不可见

    # 2) 显式设置 sheet 维度（覆盖 openpyxl 自动计算）
    dim_ref = f"A1:{last_col_letter}{last_row}"
    try:
        ws.dimensions = dim_ref
    except Exception:
        pass

    # 3) 回退占位符为空（可选，保留也可）
    # 写入内存
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    # 强制 UTF-8 编码（Windows 默认 GBK 会导致乱码）
    sys.stdin.reconfigure(encoding="utf-8")
    sys.stdout.reconfigure(encoding="utf-8")

    try:
        data = json.loads(sys.stdin.read())
        rows = data.get("rows", [])
        if not rows:
            sys.stderr.write("ERROR: empty rows\n")
            sys.exit(1)

        xlsx_bytes = build_excel(rows)
        sys.stdout.write(base64.b64encode(xlsx_bytes).decode("ascii"))

    except Exception as e:
        sys.stderr.write(f"ERROR: {e}\n")
        sys.exit(2)


if __name__ == "__main__":
    main()
